from openpyxl import load_workbook
from datetime import date
fn = 'date.xlsx'
wb = load_workbook(fn)
ws = wb['user']

def user(chat_id,first_name,phone_number):
    ws.append([chat_id,first_name,phone_number,0,0,date.today()])
    wb.save(fn)
    wb.close()

def get_cashback(message):
    w = ''
    for i in range(2,ws.max_row+1):
        if ws[f'A{i}'].value != message:
            return ws[f'D{i}'].value 
        else:
            return 'Нет такого пользователя'
      
